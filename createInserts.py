import openpyxl

# End result
result = ''

# Open workbook
wb = openpyxl.load_workbook('inserts.xlsx')

# Get the information in all sheets
for name in wb.sheetnames:
  sheet = wb[name]
  types = [] # Stores the data types

  # Count the number of elements
  i = 0
  while True:
    i += 1
    cell = sheet.cell(2, i)
    if cell.value != None:
      types.append(cell.value)
    else:
      break

  if i > 1:
    # Add the inserts
    i = 3
    while True:
      cell = sheet.cell(i, 1)
      if cell.value == None:
        break

      result += 'INSERT INTO ' + name + ' VALUES('

      # Get the values
      for j in range(len(types)):
        cell = sheet.cell(i, j + 1)
        # Since this script's main focus is Oracle SQL, the types here might not be compatible with other servers
        if ('int' in types[j]) or ('float' in types[j]) or ('double' in types[j]) or ('number' in types[j]) or ('dec' in types[j]) or ('real' in types[j]) or ('bool' in types[j]):
          # No need to use ''
          result += cell.value + ", "
        else:
          result += "'" + cell.value + "', "
      
      # Finish the expression
      # -1 removes the extra comma and space
      result = result[:-2] + ');\n\n'
      i += 1

print(result)